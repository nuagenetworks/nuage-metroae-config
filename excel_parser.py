#!/usr/bin/env python

import json
from jsonschema import validate, ValidationError
from openpyxl import load_workbook
import os
import sys

from nuage_metroae_config.util import get_expanded_template_name


TYPE_CAST_MAP = {
    "integer": int,
    "string": str,
    "boolean": bool,
    "number": float
}


def usage():
    print("Reads data from a XLSX file (Excel spreadsheet)")
    print("")
    print("Usage:")
    print("    " + " ".join([sys.argv[0],
                             "<xlsx_file>"]))
    print("")


def upper_case_string(value):
    return value.upper()


class ExcelParseError(Exception):
    pass


class ExcelParser(object):

    def __init__(self):
        self.settings = {
            "schema_directory": "schemas",
            "schemas": None,
            "use_schema_titles": False,
            "column_offset": 1,
            "row_offset": 4,
            "version_coords": "C1",
            "row_sections_present": True,
            "use_list_name": False,
            "default_fields_by_col": True}

        self.schemas = dict()
        self.errors = list()
        self.cell_positions = dict()

    def read_xlsx(self, xlsx_file):
        self.read_and_parse_schemas()

        self.data = dict()
        workbook = load_workbook(xlsx_file)
        for worksheet in workbook:
            self.read_version(worksheet, self.data)
            schema_name = self.get_schema_name(worksheet.title)
            schema = self.schemas[schema_name.lower()]
            self.data[schema_name.lower()] = self.read_worksheet(schema, worksheet)

        if len(self.errors) > 0:
            exc = ExcelParseError("There were errors while parsing "
                                  "spreadsheet, see errors member variable")
            exc.errors = self.errors
            raise exc

        return self.data

    def read_and_parse_schemas(self):
        if self.settings["schemas"] is not None:
            for name, schema_str in iter(self.settings["schemas"].items()):
                try:
                    self.schemas[name] = json.loads(schema_str)
                except Exception as e:
                    raise Exception("Could not parse schema: %s\n%s" % (
                        name, str(e)))
        else:
            for file_name in os.listdir(self.settings["schema_directory"]):
                if (file_name.endswith(".json")):
                    file_path = os.path.join(self.settings["schema_directory"],
                                             file_name)
                    with open(file_path, "rb") as f:
                        schema_str = f.read().decode("utf-8")

                    try:
                        self.schemas[file_name[0:-5]] = json.loads(schema_str)
                    except Exception as e:
                        raise Exception("Could not parse schema: %s\n%s" % (
                            file_name, str(e)))

    def read_version(self, worksheet, data):
        cell = worksheet[self.settings["version_coords"]]
        if cell.value is not None:
            data["version"] = cell.value[1:]

    def read_worksheet(self, schema, worksheet):
        if schema["type"] == "array":
            data = self.read_worksheet_list(schema, worksheet)
        else:
            data = self.read_worksheet_object(schema, worksheet)

        return data

    def read_worksheet_list(self, schema, worksheet):
        properties = schema["items"]["properties"]
        title_field_map, title_cast_map = self.generate_title_field_map(properties)

        fields_by_col = self.settings["default_fields_by_col"]
        if "fieldsByCol" in schema:
            fields_by_col = schema["fieldsByCol"]

        labels = self.read_labels(worksheet, title_field_map,
                                  fields_by_col=fields_by_col)

        data = list()
        entry_offset = 0
        while True:
            self.cell_positions.clear()
            entry = self.read_data_entry(worksheet, labels, entry_offset,
                                         title_cast_map,
                                         fields_by_col=fields_by_col)

            if entry != dict():
                self.validate_entry_against_schema(worksheet.title, [entry])
                data.append(entry)
                entry_offset += 1
            else:
                break

        if self.settings["use_list_name"] and data != list():
            list_name = self.get_list_name(schema)
            data = {list_name: data}

        return data

    def read_worksheet_object(self, schema, worksheet):
        properties = schema["properties"]
        title_field_map, title_cast_map = self.generate_title_field_map(properties)

        labels = self.read_labels(worksheet, title_field_map,
                                  fields_by_col=False)
        self.cell_positions.clear()
        data = self.read_data_entry(worksheet, labels, 0,
                                    title_cast_map, fields_by_col=False)
        if data != dict():
            self.validate_entry_against_schema(worksheet.title, data)

        return data

    def read_labels(self, worksheet, title_field_map, fields_by_col=False):
        labels = list()

        col = self.settings["column_offset"]
        row = self.settings["row_offset"]

        if self.settings["row_sections_present"] and fields_by_col:
            row += 1

        while True:
            cell = worksheet.cell(row=row, column=col)
            value = cell.value
            if fields_by_col:
                col += 1
            else:
                row += 1

            if value is not None:
                if value in title_field_map:
                    labels.append(title_field_map[value])
                else:
                    labels.append(None)
            else:
                break

        return labels

    def read_data_entry(self, worksheet, labels, entry_offset, title_cast_map,
                        fields_by_col=False):
        entry = dict()

        col = self.settings["column_offset"]
        row = self.settings["row_offset"]

        if fields_by_col:
            row += entry_offset + 1
            if self.settings["row_sections_present"]:
                row += 1
        else:
            col += entry_offset + 1

        for label in labels:
            cell = worksheet.cell(row=row, column=col)
            value = cell.value

            if sys.version_info[0] < 3 and type(value) == long:
                value = int(value)

            if value is not None:
                if label is not None:
                    if label.startswith("list:"):
                        list_name = label[5:]
                        if type(value) == int:
                            entry[list_name] = [value]
                        else:
                            entry[list_name] = [
                                title_cast_map[list_name](x.strip()) for x in value.split(",")]

                        self.cell_positions[list_name] = cell.coordinate
                    else:
                        if label in title_cast_map and type(value) != title_cast_map[label]:
                            value = title_cast_map[label](value)

                        entry[label] = value
                        self.cell_positions[label] = cell.coordinate
                else:
                    self.record_error(worksheet.title, cell.coordinate,
                                      "Data entry for unknown label")
            else:
                self.cell_positions[label] = cell.coordinate
            if fields_by_col:
                col += 1
            else:
                row += 1

        return entry

    def validate_entry_against_schema(self, schema_title, data):
        schema_name = self.get_schema_name(schema_title)
        schema = self.schemas[schema_name.lower()]

        try:
            validate(data, schema)
        except ValidationError as e:
            if e.validator == "required":
                props = e.schema["properties"]
                for field_name in e.schema["required"]:
                    if type(data) == list:
                        item = data[0]
                    else:
                        item = data
                    if field_name in props:
                        if field_name not in item:
                            title = props[field_name]["title"]
                            position = "??"
                            if field_name in self.cell_positions:
                                position = self.cell_positions[field_name]
                            self.record_error(schema_title, position,
                                              "Missing required field: " +
                                              title)
                    else:
                        self.record_error(schema_title, "??", e.message)
            elif e.relative_path[-1] in self.cell_positions:
                field_name = e.relative_path[-1]
                title = e.schema["title"]
                self.record_error(schema_title,
                                  self.cell_positions[field_name],
                                  "Invalid data for %s: %s" % (title,
                                                               e.message))
            else:
                self.record_error(schema_title, "??", e.message)

    def get_schema_name(self, title):
        if self.settings["use_schema_titles"]:
            return get_expanded_template_name(title)
        else:
            return title.replace(" ", "_").lower()

    def generate_title_field_map(self, properties):
        title_field_map = dict()
        title_cast_map = dict()
        for name, field in iter(properties.items()):
            if "type" in field and field["type"] == "array":
                title_field_map[field["title"]] = "list:" + name
                if "type" in field["items"]:
                    title_cast_map[name] = TYPE_CAST_MAP[field["items"]["type"]]
                elif "enum" in field["items"]:
                    title_cast_map[name] = upper_case_string  # special case of array with elements that are of enum type
            else:
                title_field_map[field["title"]] = name
                if "type" in field:
                    title_cast_map[name] = TYPE_CAST_MAP[field["type"]]
                elif "enum" in field:
                    if "case-sensitive" not in field:
                        title_cast_map[name] = upper_case_string  # special case of array with elements that are of enum type

        return title_field_map, title_cast_map

    def get_list_name(self, schema):
        if "listName" in schema:
            list_name = schema["listName"]
        else:
            if "items" in schema and "title" in schema["items"]:
                list_name = (
                    schema["items"]["title"].lower().replace(" ", "_") + "s")
            else:
                list_name = schema["title"].lower().replace(" ", "_")

        return list_name

    def record_error(self, schema_title, position, message):
        self.errors.append({"schema_title": schema_title,
                            "position": position,
                            "message": message})


def main():
    if len(sys.argv) != 2:
        usage()
        exit(1)

    xlsx_file = sys.argv[1]

    parser = ExcelParser()

    try:
        data = parser.read_xlsx(xlsx_file)
    except ExcelParseError:
        for error in parser.errors:
            print("%s %s | %s" % (error["schema_title"], error["position"],
                                  error["message"]))
        exit(1)

    print(json.dumps(data))


if __name__ == '__main__':
    main()
